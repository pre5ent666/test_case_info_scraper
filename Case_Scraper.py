from selenium import webdriver
from selenium.webdriver.common.by import By
import chromedriver_autoinstaller
from openpyxl import load_workbook
from openpyxl import Workbook

#--------------------------------------------- Input Settings -------------------------------------------------
chromedriver_autoinstaller.install()  # Comment this line if you failed to open webpage.
input_file = 'path/to/input/xlsx/file'
output_file = 'filename.xlsx'
username = 'your.account'
pw = 'password'


class Scraper():
    def __init__(self, test_case_list, output_name):
        self.tc_list_sheet = (load_workbook(str(test_case_list))).active
        self.output_name = str(output_name)
        self.wb = Workbook()

    def case_list(self):
        id_list = []
        for id in self.tc_list_sheet.iter_rows(max_col=1, values_only=True):
            if id is not None:
                id_list.append(str(id)[2:-3])
            else:
                break
        return id_list

    def scrapping(self,username,pw):
        print('Opening Jira...')
        driver = webdriver.Chrome()
        driver.get('https://matsjira.xxxxxx.com/login.jsp') # url of login page

        print('Entering the username and password')
        driver.find_element(By.ID,'login-form-username').send_keys(username)
        driver.find_element(By.ID,'login-form-password').send_keys(pw)
        driver.find_element(By.ID,'login-form-submit').click()

        self.wb.active
        self.wb.create_sheet('Not found', 0)
        self.wb.create_sheet('Detailed list', 0)

        id_list = self.case_list()[1:]
        cur_num = 0
        total_tc = len(id_list)

        # row_title = ['Update Date', 'Original_TCID', 'Assignee', 'GM REQ ID', 'Summary', 'TC Objective', 'Precondition', 'Test Steps', 'Expected', 'Frop List 1', 'Frop List 2', 'Story ID', 'Tag1']
        row_title = ['Original_TCID', 'Assignee', 'Precondition', 'Test Steps', 'Expected', 'TC Objective', 'Frop List 1', 'Frop List 2']
        self.wb['Detailed list'].append(row_title)

        for id in id_list:
            cur_num += 1
            print('fatching...{}/{}'.format(cur_num, total_tc))
            driver.get(self.url_gen(id))
            try:
                original_TCID = driver.find_element(By.CLASS_NAME,
                    'customfield_10202').text
                objective = driver.find_element(By.CLASS_NAME,
                    'customfield_10336').text
                precondition = driver.find_element(By.CLASS_NAME,
                    'customfield_10331').text
                test_step = driver.find_element(By.CLASS_NAME,
                    'customfield_10342').text
                expected = driver.find_element(By.CLASS_NAME,
                    'customfield_10315').text
                Frop_List_1 = driver.find_element(By.CLASS_NAME,
                    'customfield_10200').text
                Frop_List_2 = driver.find_element(By.CLASS_NAME,
                    'customfield_10319').text
                print('Found!')
                print('==========================================')
                Assignee = ''
                case_detail = [original_TCID, Assignee, precondition, test_step, expected, objective, Frop_List_1, Frop_List_2]
                self.wb['Detailed list'].append(case_detail)
            except:
                print('cannot find the detail of case: {}'.format(id))
                print('==========================================')
                self.wb['Not found'].append([id])

        print('Done!, saving the file named {}'.format(self.output_name))
        self.wb.save(self.output_name)

    def url_gen(self, tcid):
        frame = f'https://matsjira.xxxxxxx.com/issues/?jql=xxxxxxxxx' # url of jira page which containing certain test case infos
        return frame + str(tcid)

scrp = Scraper(input_file, output_file)
scrp.scrapping(username,pw)